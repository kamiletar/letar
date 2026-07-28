# Генератор XLSX-таблицы для ревью вопросов психологом.
#
# Читает батчи prisma/question-batches/<batch>/*.json и собирает
# docs/question-review/question-review-<batch>.xlsx: один лист со всеми вопросами,
# автофильтр, дропдаун вердикта (ОК/Править/Удалить), колонка комментария.
#
# Запуск (из корня apps/archetest):
#   python scripts/generate-question-review-xlsx.py           # по умолчанию 5.1
#   python scripts/generate-question-review-xlsx.py --batch 5.5
#
# Требует: pip install openpyxl

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

APP_DIR = Path(__file__).resolve().parent.parent
OUT_DIR = APP_DIR / "docs" / "question-review"

# Конфигурация батчей: порядок и русские UI-названия шкал (без клинических ярлыков)
BATCH_SCALES = {
    "5.1": [
        ("MAC", "Макиавеллизм"),
        ("HUM", "Гуманизм"),
        ("KAN", "Кантианство"),
        ("FAI", "Вера в человечество"),
        ("SAD", "Садизм"),
        ("MAS", "Мазохизм (бета)"),
        ("ASD", "Систематизация и спектр"),
        ("DIR", "Прямота"),
        ("ALX", "Алекситимия"),
    ],
    "5.5": [
        ("RES_PHYS", "Физическая броня (бета)"),
        ("RES_AFF", "Аффективный резонанс (бета)"),
        ("SPEC_INT", "Специальные интересы (бета)"),
    ],
    "hh": [
        ("HON", "Доброжелательность"),
    ],
}

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="2F5496")
REVERSE_FILL = PatternFill("solid", start_color="FFF2CC")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_scoring(scoring: dict) -> str:
    return ", ".join(f"{code} {score}" for code, score in scoring.items())


def build_questions_sheet(ws, rows):
    headers = [
        "№",
        "Шкала",
        "Код",
        "Обр.",
        "Сценарий",
        "Вариант А",
        "Вариант Б",
        "Вариант В",
        "Вариант Г",
        "Вердикт",
        "Комментарий / правка",
    ]
    widths = [5, 16, 7, 6, 45, 38, 38, 38, 38, 12, 45]

    ws.append(headers)
    for col, width in enumerate(widths, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        is_reverse = row_cells[3].value == "да"
        for cell in row_cells:
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if is_reverse:
                # Подсветка reverse-вопросов: их ключевой вариант «обратный» шкале
                if cell.column in (4,):
                    cell.fill = REVERSE_FILL

    # Дропдаун вердикта
    dv = DataValidation(type="list", formula1='"ОК,Править,Удалить"', allow_blank=True, showDropDown=False)
    dv.error = "Выберите: ОК, Править или Удалить"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{ws.max_row}")

    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    ws.freeze_panes = "F2"


INTRO_5_1 = [
    ("Ревью вопросов Архетеста — 135 новых вопросов этапа 5.1 (9 шкал × 15)", True),
    ("", False),
    ("Полная инструкция — в файле INSTRUCTIONS.md рядом с этой таблицей (или спросите у Ками).", False),
    ("Коротко:", True),
    ("• Лист «Вопросы»: один вопрос = одна строка. Варианты А–Г — с баллами шкал в скобках.", False),
    ("• «Обр.» = да — reverse-вопрос: согласие с «прямым» вариантом даёт балл ДРУГИМ шкалам,", False),
    ("  а целевая шкала набирается через отказ/обратный выбор. Проверьте, что инверсия честная.", False),
    ("• Колонка «Вердикт»: ОК / Править / Удалить (дропдаун).", False),
    ("• «Комментарий / правка»: для «Править» — предложите свою формулировку прямо в ячейке.", False),
    ("", False),
    ("Что проверять (критерии — подробнее в INSTRUCTIONS.md):", True),
    ("1. Конструкт: измеряет ли вопрос целевую шкалу; правдоподобны ли побочные баллы других шкал.", False),
    ("2. Язык: естественность, однозначность, один смысловой фокус на вариант.", False),
    ("3. Социальная желательность: «тёмные» варианты не должны выглядеть очевидно «плохими».", False),
    ("4. Этика: без патологизирующих ярлыков; вопрос не триггерный без необходимости.", False),
    ("", False),
    ("Отдельная просьба (клиническая): вопрос-маркер суицидального риска — см. раздел в INSTRUCTIONS.md.", True),
]

INTRO_5_5 = [
    ("Ревью вопросов Архетеста — 30 экспериментальных вопросов этапа 5.5 (3 шкалы × 10)", True),
    ("", False),
    ("⚠️ Это ЭКСПЕРИМЕНТАЛЬНЫЕ, авторские/прототипные шкалы ВНЕ основного ядра из 22 шкал.", False),
    ("Они видны только психологу в кабинете, всегда помечены «бета», не входят в экспресс-тест", False),
    ("и не участвуют в «ведущих чертах» пользователя. Ставка ревью — сам конструкт, а не полировка.", False),
    ("", False),
    ("Три шкалы:", True),
    ("• RES_PHYS «Физическая броня» — переносимость телесного дискомфорта/боли, приглушённость", False),
    ("  телесных сигналов (авторский конструкт, валидированного прототипа нет).", False),
    ("• RES_AFF «Аффективный резонанс» — со-переживание чужих состояний вплоть до личного дистресса", False),
    ("  (ориентир: IRI Personal Distress, Davis 1980 / HSP Scale, Aron & Aron 1997 — на уровне конструкта).", False),
    ("• SPEC_INT «Специальные интересы» — глубина погружения в тему, монополия интереса над", False),
    ("  другими сферами (авторский конструкт; ориентир — клинические описания special interests).", False),
    ("", False),
    ("Коротко по таблице:", True),
    ("• Лист «Вопросы»: один вопрос = одна строка. Варианты А–Г — с баллами шкал в скобках.", False),
    ("• Варианты скорят ТОЛЬКО экспериментальные шкалы (RES_PHYS/RES_AFF/SPEC_INT) — это сделано", False),
    ("  намеренно, чтобы не сдвигать нормализацию основного теста. Пустые скобки = вариант-дистрактор.", False),
    ("• «Обр.» = да — reverse-вопрос: высокий балл шкалы набирается «неудобным» вариантом", False),
    ("  (напр. игнорирую сигналы тела / хочется сбежать от чужого плача). Проверьте честность инверсии.", False),
    ("• «Вердикт»: ОК / Править / Удалить. «Комментарий / правка» — формулировку можно прямо в ячейке.", False),
    ("", False),
    ("Главный вопрос ревью:", True),
    ("Состоятелен ли сам конструкт? Если шкала кажется несостоятельной целиком — напишите об этом", False),
    ("отдельно, это важнее правок отдельных вопросов. Критерии — в INSTRUCTIONS.md (раздел 5.5).", False),
]

INTRO_HH = [
    ("Ревью вопросов Архетеста — 15 вопросов новой шкалы «Доброжелательность» (HON)", True),
    ("", False),
    ("⚠️ Этих вопросов ЕЩЁ НЕТ в тесте. В отличие от пакетов 5.1 и 5.5, где вопросы уже", False),
    ("работают и ревью их улучшает, здесь ваши вердикты решают, что попадёт в тест.", False),
    ("Поэтому у этого пакета есть срок: вердикты нужны к вс 02.08.", False),
    ("", False),
    ("Зачем шкала:", True),
    ("Прототип конструкта — Honesty-Humility из модели HEXACO (Ashton & Lee, 2009).", False),
    ("Формулировки авторские: пункты HEXACO не заимствуются, только уровень конструкта.", False),
    ("Практическая цель — проверить на своих данных спор в литературе: отличим ли", False),
    ("«тёмный фактор» личности от обычного антагонизма/недоброжелательности. Сейчас", False),
    ("проверить это невозможно — шкалы Доброжелательности в тесте нет.", False),
    ("", False),
    ("Главный риск, ради которого мы и просим ревью:", True),
    ("Три из четырёх фасетов Доброжелательности — почти антонимы наших тёмных шкал", False),
    ("(Искренность ↔ макиавеллизм, Справедливость ↔ антисоциальность, Скромность ↔", False),
    ("нарциссизм). Если вопросы получились ЗЕРКАЛОМ тёмных — «то же самое наоборот», —", False),
    ("то корреляция выйдет артефактом дизайна, и вся проверка обесценится.", False),
    ("Просьба смотреть каждый вопрос через этот фильтр: не является ли он просто", False),
    ("перевёрнутым вопросом про макиавеллизм/антисоциальность/нарциссизм?", False),
    ("Фасет «Нежадность» — якорный: он дальше всех от тёмных шкал (деньги, статус,", False),
    ("роскошь), и именно он даёт расхождению шанс проявиться. К нему особое внимание.", False),
    ("", False),
    ("Четыре фасета (колонка «Шкала» показывает фасет каждого вопроса):", True),
    ("• Искренность — 4 вопроса: отсутствие притворства и удобной лжи, в т. ч. без выгоды.", False),
    ("• Справедливость — 4: не пользоваться другими и правилами в свою пользу.", False),
    ("  ⚠️ Намеренно НЕ про нарушение закона — это была бы антисоциальность.", False),
    ("• Нежадность — 4: деньги, статус и роскошь не в центре жизни (якорный фасет).", False),
    ("• Скромность — 3: не претендую на особое отношение.", False),
    ("  ⚠️ Намеренно НЕ про самооценку — это была бы нарциссическая шкала.", False),
    ("", False),
    ("Коротко по таблице:", True),
    ("• Лист «Вопросы»: один вопрос = одна строка. Варианты А–Г — с баллами в скобках.", False),
    ("• Варианты скорят ТОЛЬКО HON, ни одного балла другим шкалам — это методологическое", False),
    ("  требование: общие пункты сделали бы корреляцию артефактом. Пустые скобки = дистрактор.", False),
    ("• «Обр.» = да — reverse-вопрос: балл набирается вариантом, который выглядит невыгодно", False),
    ("  или неловко (признать, что не смотрел; заплатить, хотя можно не платить; не пойти", False),
    ("  «по знакомству» ради близкого). Проверьте, что инверсия психологически честная.", False),
    ("• «Вердикт»: ОК / Править / Удалить. «Комментарий / правка» — формулировку можно", False),
    ("  прямо в ячейке. Правки внесём мы, присылать файл целиком не нужно.", False),
    ("", False),
    ("Что важнее полировки:", True),
    ("1. Зеркальность (см. выше) — главный критерий.", False),
    ("2. Социальная желательность. У этой шкалы она острее, чем у тёмных: «правильный»", False),
    ("   вариант виден. Если вопрос читается как «выберите добродетель» — он бесполезен,", False),
    ("   его выберут все. Скажите, где это так.", False),
    ("3. Покрытие фасета: 3–4 вопроса на фасет — достаточно ли они разные по ситуациям?", False),
    ("4. Язык и этика — те же критерии, что для 5.1 (см. INSTRUCTIONS.md).", False),
]

INTRO_BY_BATCH = {"5.1": INTRO_5_1, "5.5": INTRO_5_5, "hh": INTRO_HH}


def build_intro_sheet(ws, lines):
    ws.column_dimensions["A"].width = 110
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT, size=11 if bold else 10, bold=bold)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", default="5.1", choices=sorted(BATCH_SCALES))
    args = parser.parse_args()

    batch = args.batch
    scales = BATCH_SCALES[batch]
    batch_dir = APP_DIR / "prisma" / "question-batches" / batch
    out_file = OUT_DIR / f"question-review-{batch}.xlsx"

    rows = []
    num = 0
    for code, label in scales:
        batch_json = json.loads((batch_dir / f"{code}.json").read_text(encoding="utf-8"))
        for q in batch_json:
            num += 1
            options = q["options"]
            option_cells = [f"{opt['text']}\n[{fmt_scoring(opt['scoring'])}]" for opt in options]
            # Выравниваем до 4 колонок на случай нестандартного числа вариантов
            option_cells += [""] * (4 - len(option_cells))
            # Для шкал с фасетами (hh) в колонке «Шкала» показываем фасет вопроса —
            # ревьюеру важно видеть, какую грань конструкта проверяет каждый вопрос
            rows.append([
                num,
                q.get("_facet", label),
                code,
                "да" if q.get("_reverse") else "",
                q["scenario"],
                *option_cells[:4],
                "",
                "",
            ])

    wb = Workbook()
    intro = wb.active
    intro.title = "Инструкция"
    build_intro_sheet(intro, INTRO_BY_BATCH[batch])

    questions = wb.create_sheet("Вопросы")
    build_questions_sheet(questions, rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out_file)
    print(f"OK: {out_file} — {num} вопросов")


if __name__ == "__main__":
    main()
